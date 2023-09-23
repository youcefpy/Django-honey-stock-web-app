from django.shortcuts import render, redirect
from django.shortcuts import get_object_or_404
from .models import Jar, HoneyProduct, FilledJar,ProductBatch,JarBatch,Ticket,TicketBatch,Box,BoxBatch,FilledBox,FilledBoxJars,SoldFilledJar,Sku,SoldFilledBox
from django.db import transaction
from .forms import HoneyProductForm, ProductBatchForm,JarForm, JarBatchForm, FilledJarForm,TicketForm,TicketBatchForm,BoxForm,BoxBatchForm,FilledBoxForm,SkuForm,SoldFilledJarForm,ProductBatchUpdateForm,SignUpForm,SoldFilledBoxForm
from django.db.models import Sum
from .constants import BOX_CONFIG, BOX_STANDARD,BOX_TYPE_MAPPING
from django.db import IntegrityError
from django.contrib import messages
from decimal import Decimal
from django.contrib.auth.decorators import login_required
from django.http import HttpResponseForbidden
from django.contrib.auth import login
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from django.http import HttpResponse
from openpyxl.styles import Alignment,Font
from django.http import JsonResponse
import json
from .integrations.ecomanager import increment_quantity_in_ecomanager

def signup(request):
    if request.method == 'POST':
        form = SignUpForm(request.POST)
        if form.is_valid():
            user = form.save()
            # Log the user in.
            login(request, user)
            return redirect('login')  # or your desired redirect page
    else:
        form = SignUpForm()
    
    context = {
        'form' :form,
        'current_page': 'signup',
    }
    return render(request, 'signup.html', context)

@login_required
def add_honey_product(request):
    if request.method == "POST":
        form=HoneyProductForm(request.POST)
        if form.is_valid():   
            name_from_form = form.cleaned_data['name_product']
            if HoneyProduct.objects.filter(name_product=name_from_form).exists():
                messages.error(request,'Le produit exist deja dans la base de donne.')
            else : 
                form.save()
                # messages.success(request, 'Product added successfully!')
                return redirect('add_product_batch')          
    else:
        form = HoneyProductForm()
    context={
        'form':form,
    }
    return render(request,'add_honey_product.html',context)

@login_required
def add_product_batch(request):
    if request.method == "POST":
        form = ProductBatchForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('list_products')
    else : 
        form = ProductBatchForm()
    context = {
        'form':form,
    }
    return render(request,'add_product_batch.html',context)

@login_required
def list_products(request):
    products = HoneyProduct.objects.all()
    return render(request, 'list_products.html', {'products': products})

@login_required
def delete_product(request,product_id):
    if not request.user.is_staff:  # Check if the user is staff (i.e., admin)
        return HttpResponseForbidden("You don't have permission to access this page.")
    if request.method == "POST":
        product = get_object_or_404(HoneyProduct,id = product_id)
        product.delete()
        return redirect('list_products')

@login_required
def update_product_batch(request,batch_id):
    batch = get_object_or_404(ProductBatch,id=batch_id)
    if request.method == "POST":
        form = ProductBatchUpdateForm(request.POST,instance = batch)

        if form.is_valid():
            form.save()
            return redirect('list_products')
    
    else : 
        form = ProductBatchUpdateForm(instance=batch)
    
    context = {
        'form': form,
        'batch_id': batch_id,
    }
    return render(request,'update_batch_product.html',context)


@login_required
@transaction.atomic
def delete_batch(request, batch_id):
    batch = get_object_or_404(ProductBatch, id=batch_id)
    product = batch.product

    
    total_cost_before = product.weighted_average_price * product.quantity_in_the_stock
    cost_of_this_batch = batch.price_per_kg * batch.quantity_received

    
    product.quantity_in_the_stock -= batch.quantity_received

    
    new_total_quantity = product.quantity_in_the_stock
    new_total_cost = total_cost_before - cost_of_this_batch
    
    if new_total_quantity > 0:
        new_weighted_avg = new_total_cost / new_total_quantity
    else:
        new_weighted_avg = 0
    
    batch.delete()
    return redirect('list_products')


@login_required
def add_jar(request):
    if request.method == 'POST':
        form = JarForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('jars_batch')
    else:
        form = JarForm()
    return render(request, 'add_jar.html', {'form': form})


@login_required
def add_jar_batch(request):
    if request.method == "POST":
        form = JarBatchForm(request.POST)

        if form.is_valid():
            form.save()
            return redirect('list_jars')
    else :
        form = JarBatchForm()
    return render(request,'add_jar_batch.html',{'form':form})

@login_required
@transaction.atomic
def delete_jar_batch(request,batch_id):

    batch = get_object_or_404(JarBatch, id=batch_id)
    jar = batch.jar

    total_cost_before = jar.weighted_average_price * jar.quantity_in_the_stock
    cost_of_this_batch = batch.price_jar * batch.quantity_received

    
    jar.quantity_in_the_stock -= batch.quantity_received

    
    new_total_quantity = jar.quantity_in_the_stock
    new_total_cost = total_cost_before - cost_of_this_batch
    
    if new_total_quantity > 0:
        new_weighted_avg = new_total_cost / new_total_quantity
    else:
        new_weighted_avg = 0
    
    batch.delete()
    jar.save()
    return redirect('list_jars')

@login_required
def add_filled_jar(request):
    if request.method == "POST":
        form = FilledJarForm(request.POST)

        if form.is_valid():
            jar = form.cleaned_data.get('jar')
            product = form.cleaned_data.get('product')
            quantity = form.cleaned_data.get('quantity_field')
            sku = form.cleaned_data.get('sku')

            filled_jar = FilledJar.objects.filter(jar=jar, product=product).first()

            if filled_jar:
                filled_jar.quantity_field += quantity
                try:
                    filled_jar.save(update_stock=True)
                    increment_quantity_in_ecomanager(filled_jar.sku, quantity)
                except ValueError as e:
                    messages.error(request, str(e))
                    return render(request, 'add_filled_jar.html', {'form': form})
                
                # Fetch the object from the database again and print
                updated_filled_jar = FilledJar.objects.get(id=filled_jar.id)
                print(f"Updated object from DB: {updated_filled_jar.quantity_field}")

            else:
                filled_jar = FilledJar(
                    jar=jar,
                    product=product,
                    quantity_field=quantity,
                    size=jar.size,
                    sku=sku
                )
                try:
                    filled_jar.save(update_stock=True)
                    increment_quantity_in_ecomanager(filled_jar.sku, quantity)

                except ValueError as e:   # Changed IntegrityError to ValueError
                    messages.error(request, str(e))
                    return render(request, 'add_filled_jar.html', {'form': form})

            return redirect('list_filled_jars')
        
        # For form errors
        for error in form.errors.values():
            messages.error(request, error)
    else:
        form = FilledJarForm()

    return render(request, 'add_filled_jar.html', {'form': form})

@login_required
def delete_filled_jar(request, filled_jar_id):
    if not request.user.is_staff:  # Check if the user is staff (i.e., admin)
        return HttpResponseForbidden("You don't have permission to access this page.")
    
    if request.method == "POST":
        print(f"Attempting to delete filled jar with ID: {filled_jar_id}")
        filled_jar = get_object_or_404(FilledJar, id=filled_jar_id)
        try:
            with transaction.atomic():
                # Calculate honey to be added back
                honey_to_add = Decimal(filled_jar.jar.size) * Decimal(filled_jar.quantity_field)

                # Add back honey to HoneyProduct
                filled_jar.product.quantity_in_the_stock += honey_to_add
                filled_jar.product.save()
                print("Honey added back to HoneyProduct.")
            
                # Add back jars to Jar
                filled_jar.jar.quantity_in_the_stock += filled_jar.quantity_field
                filled_jar.jar.save()
                print("Jars added back.")

                # Add back the tickets
                ticket_type = f'{int(Decimal(filled_jar.jar.size) * 1000)}g'
                ticket_name = filled_jar.product.name_product
                try:
                    ticket = Ticket.objects.get(type_ticket=ticket_type, product__name_product=ticket_name)
                    ticket.quantity_in_the_stock += filled_jar.quantity_field
                    ticket.save()
                    print("Tickets added back.")
                except Ticket.DoesNotExist:
                    raise Exception(f"No ticket found for jar of size: {filled_jar.jar.size} and product: {filled_jar.product.name_product}")

                # Delete the FilledJar object
                filled_jar.delete()
                print("FilledJar object deleted.")

        except IntegrityError:
            messages.error(request, "Database integrity error.")
        except Exception as e:
            messages.error(request, f"Error deleting filled jar: {str(e)}")

        return redirect('list_filled_jars')

@login_required
def list_jars(request):
    jars = Jar.objects.all()
    return render(request, 'list_jars.html', {'jars': jars})

@login_required
def list_jars_filled(request):
    filled_jars = FilledJar.objects.all()
    aggregated_jars = FilledJar.objects.values('jar__size', 'product__name_product').annotate(total_jars=Sum('quantity_field')).order_by('jar__size', 'product__name_product')
    context = {
        'filled_jars':filled_jars,
        'aggregated_jars':aggregated_jars,
    }
    return render(request, 'list_filled_jars.html', context)


@login_required
def delete_jar(request, jar_id):
    if not request.user.is_staff:  # Check if the user is staff (i.e., admin)
        return HttpResponseForbidden("You don't have permission to access this page.")
    
    if request.method == "POST":
        jar = get_object_or_404(Jar, id=jar_id)
        jar.delete()
        return redirect('list_jars')


@login_required
def add_ticket(request):
    if request.method == "POST":
        form = TicketForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('add_batch_ticket')

    else :
        form = TicketForm()
    context = {
        'form':form
        }
    return render(request,'add_ticket.html',context)

@login_required
def add_batch_ticket(request):
    if request.method == "POST":
        form = TicketBatchForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('list_tickets')
    else : 
        form = TicketBatchForm()
    return render(request,'add_ticket_batch.html',{'form':form})

@login_required
def list_tickets(request):
    tickets = Ticket.objects.all()
    return render(request,'list_tickets.html',{'tickets':tickets})


@login_required
@transaction.atomic
def delete_ticket(request,ticket_id):
    if not request.user.is_staff:  # Check if the user is staff (i.e., admin)
        return HttpResponseForbidden("You don't have permission to access this page.")
    
    ticket=get_object_or_404(Ticket,id=ticket_id)
    ticket.delete()
    return redirect('list_tickets')


@login_required
@transaction.atomic
def delete_ticket_batch(request, batch_id):
    batch = get_object_or_404(TicketBatch, id=batch_id)
    ticket = batch.ticket

    total_cost_before = ticket.weighted_average * ticket.quantity_in_the_stock
    cost_of_this_batch = batch.purchase_price * batch.quantity_received

    ticket.quantity_in_the_stock -= batch.quantity_received

    new_total_ticket = ticket.quantity_in_the_stock
    new_total_cost = total_cost_before - cost_of_this_batch

    if new_total_ticket > 0:
        new_weighted_avg = new_total_cost / new_total_ticket
    else:
        new_weighted_avg = 0

    batch.delete()

    return redirect('list_tickets')

@login_required
def add_box(request):
    if request.method == "POST":
        form = BoxForm(request.POST)

        if form.is_valid():
            form.save()

            return redirect('add_box_batch')
    
    else :
        form = BoxForm()
    context = {
        'form':form
    }
    return render(request,'add_box.html',context)

@login_required
def add_box_batch(request):
    if request.method =='POST':
        form = BoxBatchForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('list_box')
    else : 
        form =BoxBatchForm()
    context = {
        'form':form
    }
    return render(request,'add_box_batch.html',context)

@login_required
def list_box(request):
    boxs = Box.objects.all()
    return render(request,'list_box.html',{'boxs':boxs})



@login_required
@transaction.atomic
def delete_box(request,box_id):
    if not request.user.is_staff:  # Check if the user is staff (i.e., admin)
        return HttpResponseForbidden("You don't have permission to access this page.")
    
    box=get_object_or_404(Box,id=box_id)
    box.delete()
    return redirect('list_box')


@login_required
@transaction.atomic
def delete_box_batch(request, batch_id):
    batch = get_object_or_404(BoxBatch, id=batch_id)
    box = batch.box

    total_cost_before = box.weighted_average * box.quantity_in_stock
    cost_of_this_batch = batch.purchase_price * batch.quantity_received

    box.quantity_in_stock -= batch.quantity_received

    new_total_box = box.quantity_in_stock
    new_total_cost = total_cost_before - cost_of_this_batch

    if new_total_box > 0:
        new_weighted_avg = new_total_cost / new_total_box
    else:
        new_weighted_avg = 0

    batch.delete()

    return redirect('list_box')


@login_required
def add_fill_box(request):
    
    jars_dict ={}
    if request.method == 'POST':
        form = FilledBoxForm(request.POST)
        if form.is_valid():   
            print(request.POST)   
            sku_for_filled_box = form.cleaned_data['sku_code']  
            # Extract required data
            box_type = form.cleaned_data['box_type']
            box_quantity = int(form.cleaned_data['quantity_fill_box'])
            filled_jars_data = []
            # Check if a standard is selected
            selected_standard = request.POST.get('selected_standard')

            # If a standard is selected, fetch data from BOX_STANDARD
            if selected_standard :
                box_type_id = request.POST.get('box_type') 
                box_type_str = BOX_TYPE_MAPPING.get(str(box_type_id))
                print(f"Received box_type from form: {box_type}")
                print(f"box_type_str after mapping: {box_type_str}")
                print(f"selected standard declanched : {selected_standard}")            
                selected_data = BOX_STANDARD.get(box_type_str, {}).get(selected_standard, [])
                print(f"BOX_STANDARD: {BOX_STANDARD}")
                print(f"BOX_STANDARD for box_type_str {box_type_str}: {BOX_STANDARD.get(box_type_str, {})}")
                print(f"Selected data for {box_type_str} and {selected_standard}: {selected_data}")
                filled_jars_data = [
                    (product_name, size, 1 * box_quantity)
                    for size, product_name in selected_data
                ]
                print(f"Filled jars Data  : {filled_jars_data}")
            else : 
                for key, value in request.POST.items():
                    if "Miel" in key:
                        try:
                            # Convert the value to integer and fetch the jar by its ID
                            jar_id = int(value)
                            selected_jar = FilledJar.objects.get(pk=jar_id)
                                
                            product_name = selected_jar.product.name_product
                            jar_size = selected_jar.size
                                
                            # Calculate the total jars needed
                            total_jars_needed = 1 * box_quantity

                            if (product_name,jar_size) in jars_dict:
                                jars_dict[(product_name, jar_size)] += total_jars_needed
                            else : 
                                jars_dict[(product_name,jar_size)] = total_jars_needed

                            # filled_jars_data.append((product_name, jar_size, total_jars_needed))
                           
                                
                        except (ValueError):
                            continue  
                filled_jars_data = [(key[0], key[1], value) for key, value in jars_dict.items()]
                print(f" filled jars data : ==> {filled_jars_data}")


            try: 
                # Fill the box with the provided data
                filled_box = FilledBox(box_type=box_type)
                filled_box.fill(filled_jars_data, box_quantity)

                # Update the quantity in eco_manager system using the SKU
                increment_quantity_in_ecomanager(sku_for_filled_box, box_quantity)
            except ValueError as e:
                messages.error(request, str(e))
                return redirect('add_fill_box')

            return redirect('list_filled_jars')
    else:
        form = FilledBoxForm()
    context = {
        'form': form,
        'BOX_STANDARD_JSON': json.dumps(BOX_STANDARD) ,
        'BOX_TYPE_MAPPING_JSON': json.dumps(BOX_TYPE_MAPPING) 
    }

    return render(request, 'add_fill_box.html',context)


@login_required
def list_filled_boxes(request):
    context = {}
    result = []

    all_boxes = FilledBox.objects.all().prefetch_related('filledboxjars_set__jar__product')

    box_aggregator = {}

    for box in all_boxes:
        # Create a set to ensure unique jars, so no repetition.
        unique_jars = set([f"{item.jar.product.name_product} ({item.jar.size}KG)" for item in box.filledboxjars_set.all()])
        jars_inside = " ".join(sorted(unique_jars))

        signature = f"{box.box_type.id}-{jars_inside}"

        if signature not in box_aggregator:
            box_aggregator[signature] = {
                'box_id': box.pk,
                'box_type': box.box_type,
                'jars_inside': jars_inside,
                'quantity': 0
            }

        box_aggregator[signature]['quantity'] += box.quantity_fill_box

    result = list(box_aggregator.values())

    context = {
        'result': result,
    }
    return render(request, 'list_filled_boxes.html', context)


@login_required
def delete_filled_box(request, fill_box_id):
    if not request.user.is_staff:  # Check if the user is staff (i.e., admin)
        return HttpResponseForbidden("You don't have permission to access this page.")
    """
        When we delete the filled box, we will increment the quantity of filled jars 
        back to the FilledJar and also increment the box's quantity.
    """
    filled_box = get_object_or_404(FilledBox, id=fill_box_id)
    
    # Reconstruct the filled_jars_data based on the filled box
    filled_jars_data = [
        (item.jar.product.name_product, item.jar.size, item.quantity * filled_box.quantity_fill_box) 
        for item in filled_box.filledboxjars_set.all()
    ]

    # Re-add jars to the FilledJar stock
    for product_name, jar_size, total_jars_deleted in filled_jars_data:
        try:
            filled_jar = FilledJar.objects.get(jar__size=jar_size, product__name_product=product_name)
            filled_jar.quantity_field += total_jars_deleted
            filled_jar.save()
        except FilledJar.DoesNotExist:
            messages.error(request, f"No {product_name} jars of size {jar_size} KG in stock. Delete operation aborted.")
            return redirect('list_filled_boxes')

    # Increment the box's quantity in stock
    filled_box.box_type.quantity_in_stock += filled_box.quantity_fill_box
    filled_box.box_type.save()
    
    filled_box.delete()
    
    messages.success(request, "Filled Box deleted successfully!")
    return redirect('list_filled_boxes')


@login_required
def main(request):
    products = HoneyProduct.objects.all()
    jars = Jar.objects.all()
    filledJars = FilledJar.objects.all()
    boxes = Box.objects.all()
    boxesFilled = FilledBox.objects.all()
    tickets = Ticket.objects.all()
    sell_honey_jars = SoldFilledJar.objects.all()
    sell_filled_boxes = SoldFilledBox.objects.all()

    #extraction the name and th quantity of the product 
    product_name= [product.name_product for product in products]
    product_quantity = [product.quantity_in_the_stock for product in products]

    #extraction os the quantity and the size of the size 
    jar_size = [jar.size for jar in jars]
    jar_quantity = [jar.quantity_in_the_stock for jar in jars]

    #extraction os the quantity and the size of the size 
    ticket_type_prod = [f"{ticket.product.name_product} ({ticket.type_ticket})" for ticket in tickets]
    ticket_quantity = [ticket.quantity_in_the_stock for ticket in tickets]

    #extraction of the quantity and the name and the size of the filled_jars
    prod_with_jar =[f"{filledJar.product.name_product} ({filledJar.jar.size})" for filledJar in filledJars] 
    # fill_jar_size = [fillJar.jar.size for fillJar in filledJars]
    # prod_name = [filledJar.product.name_product for filledJar in filledJars]
    prod_quantity = [filledJar.quantity_field for filledJar in filledJars]
    
    #extraction of the quantity and the type of the box 
    box_type = [box.type_box for box in boxes]
    box_quantity = [box.quantity_in_stock for box in boxes]

    #extraction of the quantity and the type of filled box 
    fill_box_type = [str(boxFilled.box_type) for boxFilled in boxesFilled ]
    fill_box_quantity = [boxFilled.quantity_fill_box for boxFilled in boxesFilled]

    #extraction of the quantity and the name and the size of the selled_jars
    sell_honey_jars_type = [f"{sell_honey.sold_date.strftime('%Y-%m-%d')} {sell_honey.filled_jar.product} ({sell_honey.filled_jar.jar.size})" for sell_honey in sell_honey_jars]
    sell_honey_jars_quantity = [sell_honey.quantity_sell_jars for sell_honey in sell_honey_jars]

    sell_filled_boxes_type =[f"{box.filled_box.box_type}" for box in sell_filled_boxes]
    sell_filled_boxes_quantity = [f"{box.quantity_sell_box}" for box in sell_filled_boxes]

    context = {
        'product_names': product_name,
        'product_quantities': product_quantity,
        'jar_size' : jar_size,
        'jar_quantity':jar_quantity,
        # 'fill_jar_size' : fill_jar_size,
        # 'prod_name':prod_name,
        'prod_with_jar':prod_with_jar,
        'prod_quantity':prod_quantity,
        'box_type':box_type,
        'box_quantity':box_quantity,
        'fill_box_type':fill_box_type,
        'fill_box_quantity':fill_box_quantity,
        'ticket_type':ticket_type_prod,
        'ticket_quantity':ticket_quantity,
        'sell_honey_jars_type':sell_honey_jars_type,
        'sell_honey_jars_quantity':sell_honey_jars_quantity,
        'sell_filled_boxes_type':sell_filled_boxes_type,
        'sell_filled_boxes_quantity':sell_filled_boxes_quantity,


    }
    return render(request,'main.html',context)

@login_required
def add_sku(request):
    if request.method == "POST":
        form = SkuForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('list_sku')
        else : 
            print(form.errors)
    else :
        form = SkuForm()
    
    return render(request,'add_sku_product.html',{'form':form})

@login_required
def list_sku(request):
    list_sku = Sku.objects.all()
    context={
        "list_sku":list_sku,
    }
    return render(request,'list_sku.html',context)

@login_required
def add_sell_fill_jars(request):
    if request.method == "POST":
        form = SoldFilledJarForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('list_sell_jars')

    else:
        form = SoldFilledJarForm()
    context={
        'form':form
    }
    return render(request,'add_sell_filled_jars.html',context)

@login_required
def add_sell_fill_box(request):
    if request.method == "POST":
        form = SoldFilledBoxForm(request.POST)
        try:

            if form.is_valid():
                form.save()
                return redirect('list_sell_boxes')
        except ValueError as e:
            messages.error(request,str(e))
            
    else : 
        form = SoldFilledBoxForm()
    
    context={
        'form':form,
    }
    return render(request,'add_sell_filled_box.html',context)

@login_required
def list_sell_jars(request):
    list_sell_jars =SoldFilledJar.objects.all()
    context={
        'list_sell_jars':list_sell_jars
    }
    return render(request,'list_sell_jars.html',context)


@login_required
def list_sell_boxes(request):
    list_sell_box=SoldFilledBox.objects.all()
    context={
        'list_sell_box':list_sell_box
    }
    return render(request,'list_sell_boxes.html',context)

def export_products_to_excel(request):
    wb = Workbook()
    ws = wb.active
    ws.title = "Products"

    # Header
    columns = ['Nom Produit', 'Quantiter de stock en (kg)', 'Prix/KG', 'Historique des Achat']
    for col_num, column_title in enumerate(columns, 1):
        col_letter = get_column_letter(col_num)
        ws['{}1'.format(col_letter)] = column_title
        ws.column_dimensions[col_letter].width = 15

    # Data
    products = HoneyProduct.objects.all()
    for idx, product in enumerate(products, 2):
        ws.cell(row=idx, column=1, value=product.name_product)
        ws.cell(row=idx, column=2, value=product.quantity_in_the_stock)
        ws.cell(row=idx, column=3, value=product.weighted_average_price)
        
        batches = product.productbatch_set.all()
        batch_details = '\n'.join([f"Date: {batch.date_received}, Quantiter: {batch.quantity_received}kg, Prix: {batch.price_per_kg} DA" for batch in batches])
        ws.cell(row=idx, column=4, value=batch_details).alignment=Alignment(wrap_text=True)

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=products.xlsx'
    wb.save(response)

    return response


def export_jars_to_excel(request):
    wb = Workbook()
    ws = wb.active
    ws.title='BOCALS'

    #Header
    columns = ['Taille','Quantiter dans le stoque','Prix/bocal','Historique des Achat']
    for col_num,column_title in enumerate(columns,1):
        col_lettre = get_column_letter(col_num)
        ws['{}1'.format(col_lettre)]=column_title
        ws.column_dimensions[col_lettre].width = 15

    #Data
    jars = Jar.objects.all()
    for idx, jar in enumerate(jars,2):
        ws.cell(row=idx,column=1,value=jar.size)
        ws.cell(row=idx,column=2,value=jar.quantity_in_the_stock)
        ws.cell(row=idx,column=3,value=jar.weighted_average_price)

        batches = jar.jarbatch_set.all()
        batch_details = '\n'.join([f'Date: {batch.date_received} | Quantiter: {batch.quantity_received} | Prix: {batch.price_jar} DA' for batch in batches ])
        ws.cell(row=idx,column=4,value=batch_details).alignment=Alignment(wrap_text=True)


    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=jars.xlsx'
    wb.save(response)

    return response


def export_tickets_excel(request):
    wb = Workbook()
    ws = wb.active
    ws.title = 'Etiquette'

    #headers 
    columns=['Type','Nom de Produit','Quantiter Dans Le Stock','Prix','Historique D\'achat']
    for cul_num,column_title in enumerate(columns,1):
        col_lettre = get_column_letter(cul_num)
        ws['{}1'.format(col_lettre)] = column_title
        ws.column_dimensions[col_lettre].width=15

    #data
    tickets =Ticket.objects.all()
    for idx,ticket in enumerate(tickets,2):
        ws.cell(row=idx,column=1,value=ticket.type_ticket)
        ws.cell(row=idx,column=2,value=ticket.product.name_product)
        ws.cell(row=idx,column=3,value=ticket.quantity_in_the_stock)
        ws.cell(row=idx,column=4,value=ticket.weighted_average) 

        batches = ticket.ticketbatch_set.all()
        batch_details ='\n'.join([f"Date: {batch.date_entry} | Quantiter: {batch.quantity_received} | Prix/Etiquette: {batch.purchase_price}"for batch in batches])
        ws.cell(row=idx,column=5,value=batch_details).alignment=Alignment(wrap_text=True)


    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=tickets.xlsx'
    wb.save(response)

    return response

def export_filled_jars_excel(request):
    wb = Workbook()
    ws = wb.active
    ws.title = 'Bocal REmpli'

    #headers 
    columns=['Taille Du Bocal','Nom Produit','Quantiter des bocal rempli','Date de remplisage']
    for cul_num,column_title in enumerate(columns,1):
        col_lettre = get_column_letter(cul_num)
        ws['{}1'.format(col_lettre)] = column_title
        ws.column_dimensions[col_lettre].width=15

    #data
    filledJars =FilledJar.objects.all()
    for idx,filledjar in enumerate(filledJars,2):
        ws.cell(row=idx,column=1,value=filledjar.jar.size)
        ws.cell(row=idx,column=2,value=filledjar.product.name_product)
        ws.cell(row=idx,column=3,value=filledjar.quantity_field)
        ws.cell(row=idx,column=4,value=filledjar.filled_date) 

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=filled_jars.xlsx'
    wb.save(response)

    return response

def export_boxes_to_excel(request):
    # Create a new Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Coffret"

    # Add headers to the Excel file
    headers = ['Type de coffret', 'Quantiter dans le Stoque', 'Price', 'Historique des achat/lot']
    for col_num, header in enumerate(headers, 1):
        col_letter = get_column_letter(col_num)
        ws[f'{col_letter}1'] = header
        ws[f'{col_letter}1'].font = Font(bold=True)
    
    # Add data to the Excel file
    for idx, box in enumerate(Box.objects.all(), 2):
        ws.cell(row=idx, column=1, value=box.get_type_box_display())
        ws.cell(row=idx, column=2, value=box.quantity_in_stock)
        ws.cell(row=idx, column=3, value=f"{box.weighted_average:.2f} DA")
        batches = box.boxbatch_set.all()
        batch_details = "\n".join([f"Date: {batch.date_entry} | Quantiter: {batch.quantity_received} | Prix/Coffret: {batch.purchase_price} DA" for batch in batches])
        ws.cell(row=idx, column=4, value=batch_details).alignment = Alignment(wrap_text=True)

    # Serve the Excel file to the user
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=boxes.xlsx'
    wb.save(response)
    return response

def export_filled_boxes_excel(request):
    wb = Workbook()
    ws = wb.active
    ws.title = "Coffret Rempli"

    # Add headers to the Excel file
    headers = ['Type de coffret', 'Bocal Dans Coffret', 'Quantiter']
    for col_num, header in enumerate(headers, 1):
        col_letter = get_column_letter(col_num)
        ws[f'{col_letter}1'] = header
        ws[f'{col_letter}1'].font = Font(bold=True)
    
    for idx, box in enumerate(FilledBox.objects.all(), 2):
        ws.cell(row=idx, column=1, value=str(box.box_type))
        ws.cell(row=idx, column=3, value=f"{box.quantity_fill_box:.2f} DA")

    # Serve the Excel file to the user
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=coffret_rempli.xlsx'
    wb.save(response)
    return response


def sku_search(request):
    q = request.GET.get('q', '')
    skus = Sku.objects.filter(code__icontains=q).values('id', 'code')
    return JsonResponse(list(skus), safe=False)


def get_jars(request):
    
    box_type_id = request.GET.get('box_type')
    if not box_type_id:
        return JsonResponse({'error': 'Invalid box type ID'}, status=400)
    try:
        box_type_obj = Box.objects.get(pk=box_type_id)
    except Box.DoesNotExist:
        return JsonResponse({'error': 'Box type not found'}, status=404)
    box_type_value = box_type_obj.type_box

    if box_type_value in ['coffret 3500']:
        num_jars=3
        filled_jars = FilledJar.objects.filter(jar__size__lte=0.25)
    
    elif box_type_value in ['coffret 4000', 'coffret 4500']:
        num_jars = 4
        filled_jars = FilledJar.objects.filter(jar__size__lte=0.25)
    else:
        num_jars = 6
        filled_jars = FilledJar.objects.all()

    jar_data = []
    for jar in filled_jars:
        jar_data.append({
            'id': jar.pk,
            'name': jar.product.name_product,
            'size': jar.size
        })

    response_data = {
        'num_jars': num_jars,
        'filled_jars': jar_data
    }
    print(response_data)
    # print(f"Box Type Requested: {box_type}")
    print(f"Jars Returned: {filled_jars}")

    return JsonResponse(response_data)
